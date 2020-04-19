import os
from openpyxl import load_workbook, Workbook

PWD = os.path.abspath(os.path.dirname(__file__))


#region read+write xlsx util
"""
wb aka workbook
ws aka worksheet
"""

class ExcelRead:

    def openpyxl_read_excel(self, path_file):
        wb = load_workbook(path_file)
        ws = wb.active

        max_row = ws.max_row
        max_column = ws.max_column

        data_return = ()
        for i in range(1, max_row + 1):  # iterate over all cells
            row = []
            for j in range(1, max_column + 1):
                cell_obj = ws.cell(row=i, column=j)
                row.append(cell_obj.value)
            data_return += (row,)
        return data_return


class ExcelWrite:

    def openpyxl_write_new_excel_file(self, file_name, data_to_write=()):
        wb = Workbook()
        ws = wb.active

        # append all rows
        for row in data_to_write:
            ws.append(tuple(row))

        # save file
        wb.save(file_name)
    
    def openpyxl_write_existing_excel_file(self, file_name, data_to_write=()):
        wb = load_workbook(file_name)
        ws = wb.active
        max_row = ws.max_row
        for row in data_to_write:
            for i in range(1, len(row) + 1):  ## first start column and row are 1
                cell = ws.cell(row=max_row + 1, column=i)
                cell.value = row[i - 1]
            max_row += 1 ## add 1 to write new row
        wb.save(file_name)

#endregion read+write xlsx util


class Test:

    def test_openpyxl_read(self):
        # create fixture
        EXP = (['abb', 122], ['xxx', 333])
        excel_file_name = f'{PWD}/fixture/input.xlsx'
        excelread = ExcelRead()

        # check excel file is exist and it's data
        assert os.path.isfile(excel_file_name)
        assert excelread.openpyxl_read_excel(excel_file_name) == EXP

    def test_openpyxl_write_new_file(self):
        # create fixture
        expenses = (['abb', 122], ['xxx', 333])
        excel_file_name, worksheet_name = f'{PWD}/tmp/output.xlsx', 'Sheet test'
        excelread, excelwrite= ExcelRead(), ExcelWrite()
        excelwrite.openpyxl_write_new_excel_file(excel_file_name,  expenses)

        # check excel file is exist and it's data
        assert os.path.isfile(excel_file_name)
        assert excelread.openpyxl_read_excel(excel_file_name) == expenses

    def test_openpyxl_update_current_file(self):
        # create fixture
        excel_file_name, worksheet_name = f'{PWD}/tmp/output.xlsx', 'Sheet test'
        excelread, excelwrite= ExcelRead(), ExcelWrite()
        old_expenese = excelread.openpyxl_read_excel(excel_file_name)
        new_expenses = (['Telephone', 300], ['Party', 200])
        excelwrite.openpyxl_write_existing_excel_file(excel_file_name,  new_expenses) ## write new_expenses
        expenses = old_expenese + new_expenses

        # check excel file is exist and it's data
        assert os.path.isfile(excel_file_name)
        assert excelread.openpyxl_read_excel(excel_file_name) == expenses
